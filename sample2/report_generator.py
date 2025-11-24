"""
Report Generator for Feature Analysis
Creates Excel and PDF reports with plots and statistics
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import LineChart, Reference, BarChart
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from datetime import datetime
import os
import json
from typing import Dict, List, Optional
from PIL import Image
import io


class ReportGenerator:
    """
    Generate Excel and PDF reports for feature analysis
    """
    
    def __init__(self):
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="00395D", end_color="00395D", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_excel_report(self, 
                            df: pd.DataFrame,
                            output_dir: str,
                            metadata: Dict,
                            identifier: str = "") -> str:
        """
        Generate comprehensive Excel report
        
        Parameters:
        -----------
        df : pd.DataFrame
            Processed dataframe with all features
        output_dir : str
            Directory where plots are saved
        metadata : Dict
            Metadata dictionary with statistics
        identifier : str
            Account or margin center identifier
            
        Returns:
        --------
        str
            Path to generated Excel file
        """
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, metadata, identifier)
        
        # 2. Statistics Sheet
        ws_stats = wb.create_sheet("Statistics")
        self._create_statistics_sheet(ws_stats, df, metadata)
        
        # 3. Time Series Data Sheet
        ws_timeseries = wb.create_sheet("Time Series Data")
        self._create_timeseries_sheet(ws_timeseries, df)
        
        # 4. Correlation Matrix Sheet
        ws_corr = wb.create_sheet("Correlations")
        self._create_correlation_sheet(ws_corr, df)
        
        # 5. Feature Analysis Sheet
        ws_features = wb.create_sheet("Feature Analysis")
        self._create_feature_analysis_sheet(ws_features, df)
        
        # 6. Anomaly Candidates Sheet (high delta values)
        ws_anomaly = wb.create_sheet("Anomaly Candidates")
        self._create_anomaly_sheet(ws_anomaly, df)
        
        # 7. Plots Overview Sheet (with embedded images if possible)
        ws_plots = wb.create_sheet("Plots Overview")
        self._create_plots_sheet(ws_plots, output_dir)
        
        # Save workbook
        report_path = os.path.join(output_dir, f'report_{identifier}_{datetime.now():%Y%m%d}.xlsx')
        wb.save(report_path)
        
        print(f"Excel report saved to {report_path}")
        return report_path
    
    def _create_summary_sheet(self, ws, metadata: Dict, identifier: str):
        """Create summary sheet with key information"""
        ws.append(["Feature Analysis Report"])
        ws['A1'].font = Font(bold=True, size=16)
        
        ws.append([])
        ws.append(["Report Information"])
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill
        
        # Add metadata
        ws.append(["Generated:", metadata.get('generation_timestamp', '')])
        ws.append(["Identifier:", identifier])
        ws.append(["Data Period:", f"{metadata['data_period']['start']} to {metadata['data_period']['end']}"])
        ws.append(["Total Days:", metadata['data_period']['total_days']])
        ws.append(["Total Records:", metadata['data_stats']['total_records']])
        
        ws.append([])
        ws.append(["Features Analyzed"])
        ws['A10'].font = self.header_font
        ws['A10'].fill = self.header_fill
        
        ws.append(["Basic Features:", ", ".join(metadata['features_analyzed']['basic'][:5])])
        ws.append(["Computed Features:", f"{len(metadata['features_analyzed']['computed'])} features"])
        ws.append(["Engineered Features:", f"{len(metadata['features_analyzed']['engineered'])} features"])
        
        # Format column width
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
    
    def _create_statistics_sheet(self, ws, df: pd.DataFrame, metadata: Dict):
        """Create detailed statistics sheet"""
        ws.append(["Feature Statistics Summary"])
        ws['A1'].font = Font(bold=True, size=14)
        
        ws.append([])
        
        # Create header row
        headers = ["Feature", "Mean", "Std Dev", "Min", "25%", "Median", "75%", "Max", "Skewness", "Kurtosis"]
        ws.append(headers)
        
        # Style header row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add statistics for each feature
        row = 4
        for feature, stats in metadata.get('summary_statistics', {}).items():
            # Calculate additional statistics if needed
            if feature in df.columns:
                feature_data = df[feature].dropna()
                skewness = feature_data.skew() if len(feature_data) > 0 else 0
                kurtosis = feature_data.kurtosis() if len(feature_data) > 0 else 0
                q25 = feature_data.quantile(0.25) if len(feature_data) > 0 else 0
                q75 = feature_data.quantile(0.75) if len(feature_data) > 0 else 0
                
                ws.append([
                    feature,
                    f"{stats['mean']:.2f}",
                    f"{stats['std']:.2f}",
                    f"{stats['min']:.2f}",
                    f"{q25:.2f}",
                    f"{stats['median']:.2f}",
                    f"{q75:.2f}",
                    f"{stats['max']:.2f}",
                    f"{skewness:.2f}",
                    f"{kurtosis:.2f}"
                ])
                
                # Add borders
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).border = self.border
                
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_timeseries_sheet(self, ws, df: pd.DataFrame):
        """Create time series data sheet"""
        ws.append(["Time Series Data"])
        ws['A1'].font = Font(bold=True, size=14)
        
        # Select columns to include (limit to prevent Excel from being too large)
        cols_to_include = ['Business_Date'] + [col for col in df.columns 
                                               if not col.endswith('_delta') and 
                                               not col.endswith('_delta_pct') and
                                               not col.endswith('_zscore')][:10]
        
        # Create subset dataframe
        df_subset = df[cols_to_include].copy()
        
        # Convert to rows and add to sheet
        for r in dataframe_to_rows(df_subset, index=False, header=True):
            ws.append(r)
        
        # Style header row
        for cell in ws[2]:
            cell.font = self.header_font
            cell.fill = self.header_fill
    
    def _create_correlation_sheet(self, ws, df: pd.DataFrame):
        """Create correlation matrix sheet"""
        ws.append(["Feature Correlation Matrix"])
        ws['A1'].font = Font(bold=True, size=14)
        
        # Select numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns[:15]  # Limit to 15 for readability
        
        if len(numeric_cols) > 1:
            # Calculate correlation matrix
            corr_matrix = df[numeric_cols].corr()
            
            # Add to sheet
            ws.append([])
            ws.append([''] + list(corr_matrix.columns))
            
            # Style header row
            for col in range(2, len(corr_matrix.columns) + 2):
                cell = ws.cell(row=3, column=col)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center', text_rotation=90)
            
            # Add correlation values
            for idx, row_name in enumerate(corr_matrix.index):
                row_data = [row_name] + [f"{val:.2f}" for val in corr_matrix.iloc[idx]]
                ws.append(row_data)
                
                # Style row header
                ws.cell(row=4+idx, column=1).font = self.header_font
                ws.cell(row=4+idx, column=1).fill = self.header_fill
                
                # Color code correlation values
                for col_idx, val in enumerate(corr_matrix.iloc[idx]):
                    cell = ws.cell(row=4+idx, column=2+col_idx)
                    
                    # Apply color based on correlation strength
                    if abs(val) > 0.7:
                        cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    elif abs(val) > 0.5:
                        cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
                    elif abs(val) > 0.3:
                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    def _create_feature_analysis_sheet(self, ws, df: pd.DataFrame):
        """Create feature analysis sheet with delta statistics"""
        ws.append(["Feature Delta Analysis"])
        ws['A1'].font = Font(bold=True, size=14)
        
        ws.append([])
        headers = ["Feature", "Mean Delta", "Std Delta", "Max Delta", "Min Delta", 
                  "Mean Delta %", "Max Delta %", "Outliers (>3σ)"]
        ws.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        row = 4
        # Analyze delta features
        for col in df.columns:
            if col.endswith('_delta') and not col.endswith('_pct') and not col.endswith('_zscore'):
                base_feature = col.replace('_delta', '')
                delta_pct_col = f'{base_feature}_delta_pct'
                
                if delta_pct_col in df.columns:
                    delta_data = df[col].dropna()
                    delta_pct_data = df[delta_pct_col].replace([np.inf, -np.inf], np.nan).dropna()
                    
                    # Calculate outliers
                    mean_val = delta_data.mean()
                    std_val = delta_data.std()
                    outliers = (np.abs(delta_data - mean_val) > 3 * std_val).sum()
                    
                    ws.append([
                        base_feature,
                        f"{delta_data.mean():.2f}",
                        f"{delta_data.std():.2f}",
                        f"{delta_data.max():.2f}",
                        f"{delta_data.min():.2f}",
                        f"{delta_pct_data.mean():.2f}%" if len(delta_pct_data) > 0 else "N/A",
                        f"{delta_pct_data.max():.2f}%" if len(delta_pct_data) > 0 else "N/A",
                        outliers
                    ])
                    
                    # Add borders
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col_idx).border = self.border
                    
                    row += 1
    
    def _create_anomaly_sheet(self, ws, df: pd.DataFrame):
        """Create sheet with potential anomaly candidates"""
        ws.append(["Anomaly Candidates (High Delta Values)"])
        ws['A1'].font = Font(bold=True, size=14)
        
        ws.append([])
        ws.append(["Criteria: Values exceeding 3 standard deviations from mean"])
        
        ws.append([])
        headers = ["Date", "Feature", "Value", "Delta", "Z-Score", "Deviation from Mean"]
        ws.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=5, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        row = 6
        anomaly_count = 0
        max_anomalies = 100  # Limit to prevent sheet from being too large
        
        # Check each delta feature for anomalies
        for col in df.columns:
            if col.endswith('_delta') and not col.endswith('_pct') and not col.endswith('_zscore'):
                base_feature = col.replace('_delta', '')
                
                if base_feature in df.columns:
                    delta_data = df[col].dropna()
                    if len(delta_data) > 0:
                        mean_val = delta_data.mean()
                        std_val = delta_data.std()
                        
                        # Find anomalies
                        if std_val > 0:
                            z_scores = (df[col] - mean_val) / std_val
                            anomaly_mask = np.abs(z_scores) > 3
                            
                            anomaly_df = df[anomaly_mask][['Business_Date', base_feature, col]].copy()
                            anomaly_df['z_score'] = z_scores[anomaly_mask]
                            anomaly_df['deviation'] = df[col][anomaly_mask] - mean_val
                            
                            for _, anomaly_row in anomaly_df.iterrows():
                                if anomaly_count >= max_anomalies:
                                    break
                                
                                ws.append([
                                    str(anomaly_row['Business_Date']),
                                    base_feature,
                                    f"{anomaly_row[base_feature]:.2f}",
                                    f"{anomaly_row[col]:.2f}",
                                    f"{anomaly_row['z_score']:.2f}",
                                    f"{anomaly_row['deviation']:.2f}"
                                ])
                                
                                # Highlight row
                                for col_idx in range(1, len(headers) + 1):
                                    cell = ws.cell(row=row, column=col_idx)
                                    cell.border = self.border
                                    if abs(anomaly_row['z_score']) > 4:
                                        cell.fill = PatternFill(start_color="FF9999", 
                                                              end_color="FF9999", 
                                                              fill_type="solid")
                                
                                row += 1
                                anomaly_count += 1
        
        ws.append([])
        ws.append([f"Total anomaly candidates shown: {min(anomaly_count, max_anomalies)}"])
    
    def _create_plots_sheet(self, ws, output_dir: str):
        """Create sheet with plot references"""
        ws.append(["Generated Plots Overview"])
        ws['A1'].font = Font(bold=True, size=14)
        
        ws.append([])
        ws.append(["Plot Name", "Description", "File Path"])
        
        # Style headers
        for col in range(1, 4):
            cell = ws.cell(row=3, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # List all PNG files in output directory
        plot_descriptions = {
            '01_combined_overview': 'Combined overview of all features',
            '02_': 'Individual feature analysis',
            '08_correlation_heatmap': 'Feature correlation matrix',
            '09_statistical_summary': 'Statistical summary plots',
            'consolidated': 'Consolidated margin center view',
            '_engineered': 'Engineered feature analysis'
        }
        
        row = 4
        if os.path.exists(output_dir):
            for file in sorted(os.listdir(output_dir)):
                if file.endswith('.png'):
                    # Determine description
                    description = "Feature plot"
                    for key, desc in plot_descriptions.items():
                        if key in file:
                            description = desc
                            break
                    
                    ws.append([
                        file,
                        description,
                        os.path.join(output_dir, file)
                    ])
                    
                    # Add borders
                    for col_idx in range(1, 4):
                        ws.cell(row=row, column=col_idx).border = self.border
                    
                    row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
    
    def generate_pdf_report(self, 
                          output_dir: str,
                          metadata: Dict,
                          identifier: str = "") -> str:
        """
        Generate PDF report with all plots
        
        Parameters:
        -----------
        output_dir : str
            Directory containing plots
        metadata : Dict
            Metadata dictionary
        identifier : str
            Account or margin center identifier
            
        Returns:
        --------
        str
            Path to generated PDF file
        """
        pdf_path = os.path.join(output_dir, f'report_{identifier}_{datetime.now():%Y%m%d}.pdf')
        
        with PdfPages(pdf_path) as pdf:
            # Title page
            fig = plt.figure(figsize=(8.5, 11))
            fig.text(0.5, 0.7, 'Feature Analysis Report', 
                    ha='center', size=24, weight='bold')
            fig.text(0.5, 0.6, f'Identifier: {identifier}', 
                    ha='center', size=16)
            fig.text(0.5, 0.5, f"Period: {metadata['data_period']['start']} to {metadata['data_period']['end']}", 
                    ha='center', size=14)
            fig.text(0.5, 0.4, f"Generated: {datetime.now():%Y-%m-%d %H:%M}", 
                    ha='center', size=12)
            
            # Add summary statistics
            summary_text = f"""
            Total Records: {metadata['data_stats']['total_records']}
            Total Days: {metadata['data_period']['total_days']}
            Features Analyzed: {len(metadata['features_analyzed']['basic'])} basic, 
            {len(metadata['features_analyzed']['computed'])} computed,
            {len(metadata['features_analyzed']['engineered'])} engineered
            """
            fig.text(0.5, 0.2, summary_text, ha='center', size=10)
            
            plt.axis('off')
            pdf.savefig(fig, bbox_inches='tight')
            plt.close()
            
            # Add all plot images
            if os.path.exists(output_dir):
                for file in sorted(os.listdir(output_dir)):
                    if file.endswith('.png'):
                        try:
                            img_path = os.path.join(output_dir, file)
                            img = Image.open(img_path)
                            
                            # Create figure with appropriate size
                            fig = plt.figure(figsize=(11, 8.5))
                            ax = fig.add_subplot(111)
                            ax.imshow(img)
                            ax.axis('off')
                            ax.set_title(file.replace('.png', '').replace('_', ' ').title(), 
                                       fontsize=12, pad=20)
                            
                            pdf.savefig(fig, bbox_inches='tight')
                            plt.close()
                        except Exception as e:
                            print(f"Error adding {file} to PDF: {str(e)}")
            
            # Add metadata information
            d = pdf.infodict()
            d['Title'] = f'Feature Analysis Report - {identifier}'
            d['Author'] = 'Feature Plotter Service'
            d['Subject'] = 'Historical Feature Analysis'
            d['Keywords'] = 'Features, Analysis, Anomaly Detection'
            d['CreationDate'] = datetime.now()
        
        print(f"PDF report saved to {pdf_path}")
        return pdf_path


# Integration function
def generate_reports(df: pd.DataFrame, 
                    output_dir: str,
                    metadata: Dict,
                    identifier: str = "",
                    report_types: List[str] = ['excel', 'pdf']) -> Dict[str, str]:
    """
    Generate reports in specified formats
    
    Parameters:
    -----------
    df : pd.DataFrame
        Processed dataframe
    output_dir : str
        Output directory with plots
    metadata : Dict
        Metadata dictionary
    identifier : str
        Account or margin center identifier
    report_types : List[str]
        Types of reports to generate ('excel', 'pdf')
        
    Returns:
    --------
    Dict[str, str]
        Dictionary with paths to generated reports
    """
    generator = ReportGenerator()
    reports = {}
    
    if 'excel' in report_types:
        reports['excel'] = generator.generate_excel_report(df, output_dir, metadata, identifier)
    
    if 'pdf' in report_types:
        reports['pdf'] = generator.generate_pdf_report(output_dir, metadata, identifier)
    
    return reports
